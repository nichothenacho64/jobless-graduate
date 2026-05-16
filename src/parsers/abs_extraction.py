from __future__ import annotations

from collections.abc import Mapping
from typing import Optional, cast

from openpyxl import load_workbook
import pandas as pd

from src.loaders import list_excel_sheets, resolve_spreadsheet_path
from src.parsers.constants import (
    ABS_CONTENTS_SHEET_NAME,
    ABS_FOOTER_NOTE_PATTERN,
    ABS_FOOTER_PREFIXES,
    ABS_FOOTER_WORD_MARKERS,
    ABS_HEADER_FIRST_ROW_INDEX,
    ABS_MEASUREMENT_LABELS,
    ABS_TABLE_LABEL_PATTERN,
    ABS_TABLE_SOURCE_KEY_PATTERN,
    ABS_TITLE_ROW_INDEX,
)
from src.parsers.sheets import (
    clean_cell_text,
    count_nonempty_cells,
    get_row_texts,
    normalise_cell_text,
    require_cell_text,
)
from src.types import (
    ABSSourceBounds,
    ABSMeasurement,
    ABSMeasurementCell,
    ABSParsedTable,
    ABSRowBounds,
    BodyRowKind,
    Folder,
    Metadata,
)


def _last_nonblank_row(raw_sheet: pd.DataFrame) -> int:
    for row_index in range(len(raw_sheet) - 1, -1, -1):
        if count_nonempty_cells(raw_sheet.iloc[row_index]) > 0:
            return row_index

    raise ValueError("Could not identify a nonblank ABS worksheet row.")


def _row_has_measure_values(
    raw_sheet: pd.DataFrame,
    row_index: int,
    bounds: ABSSourceBounds,
) -> bool:
    for col_index in range(bounds.col_first, bounds.col_last + 1):
        if clean_cell_text(raw_sheet.iat[row_index, col_index]):
            return True

    return False


def _is_footer_row(row: pd.Series) -> bool:
    row_texts = get_row_texts(row)
    if not row_texts:
        return False

    first_text = row_texts[0].lstrip()
    first_text_lower = first_text.lower()

    return (
        first_text.startswith(ABS_FOOTER_PREFIXES)
        or first_text_lower.startswith(ABS_FOOTER_WORD_MARKERS)
        or ABS_FOOTER_NOTE_PATTERN.search(first_text_lower) is not None
    )


def _next_body_row_kind(
    raw_sheet: pd.DataFrame,
    row_bounds: ABSRowBounds,
    bounds: ABSSourceBounds,
    start_row: int,
) -> Optional[BodyRowKind]:
    for row_index in range(start_row, row_bounds.body_last + 1):
        row = raw_sheet.iloc[row_index]
        if count_nonempty_cells(row) == 0:
            continue

        row_label = clean_cell_text(raw_sheet.iat[row_index, 0])
        if not row_label:
            return None

        if _row_has_measure_values(raw_sheet, row_index, bounds):
            return "values"

        return "label"

    return None


def _next_measurement_row(
    measurement_cells: list[ABSMeasurementCell],
    measurement_row: int,
) -> Optional[int]:
    next_row: Optional[int] = None

    for cell in measurement_cells:
        if cell.row <= measurement_row:
            continue

        if next_row is None or cell.row < next_row:
            next_row = cell.row

    return next_row


def _infer_subtable_body_last(
    sheet_body_last: int,
    measurement_cell: ABSMeasurementCell,
    measurement_cells: list[ABSMeasurementCell],
) -> int:
    next_row = _next_measurement_row(measurement_cells, measurement_cell.row)
    if next_row is None:
        return sheet_body_last

    return next_row - 1


def _next_measurement_col_on_same_row(
    measurement_cell: ABSMeasurementCell,
    measurement_cells: list[ABSMeasurementCell],
) -> Optional[int]:
    next_col: Optional[int] = None

    for cell in measurement_cells:
        if cell.row != measurement_cell.row or cell.col <= measurement_cell.col:
            continue

        if next_col is None or cell.col < next_col:
            next_col = cell.col

    return next_col


def _infer_subtable_col_last(
    raw_sheet: pd.DataFrame,
    rows: ABSRowBounds,
    measurement_cell: ABSMeasurementCell,
    measurement_cells: list[ABSMeasurementCell],
    body_last: int,
) -> int:
    next_col = _next_measurement_col_on_same_row(measurement_cell, measurement_cells)
    search_stop_col = next_col - 1 if next_col is not None else raw_sheet.shape[1] - 1

    for col_index in range(search_stop_col, measurement_cell.col - 1, -1):
        for row_index in range(rows.header_first, body_last + 1):
            if clean_cell_text(raw_sheet.iat[row_index, col_index]):
                return col_index

    return measurement_cell.col


def parse_abs_table_label(text: str) -> tuple[int, str]:
    match = ABS_TABLE_LABEL_PATTERN.search(text)
    if match is None:
        raise ValueError(f"Could not parse an ABS table label from {text!r}.")

    table_number = int(match.group("NUMBER"))
    title = match.group("TITLE").strip()

    return table_number, title


def find_abs_title(raw_sheet: pd.DataFrame) -> str:
    title_extraction_error_message = "Could not extract a title from the ABS sheet."
    source_title = require_cell_text(
        raw_sheet,
        ABS_TITLE_ROW_INDEX,
        0,
        title_extraction_error_message,
    )
    _, title = parse_abs_table_label(source_title)

    return title


def parse_abs_sheet_number(sheet_name: str) -> int:
    sheet_number, _ = parse_abs_table_label(sheet_name)
    return sheet_number


def find_abs_measurement_cells(raw_sheet: pd.DataFrame) -> list[ABSMeasurementCell]:
    measurement_cells: list[ABSMeasurementCell] = []

    for row_index in range(raw_sheet.shape[0]):
        for col_index in range(raw_sheet.shape[1]):
            label = normalise_cell_text(raw_sheet.iat[row_index, col_index])
            measurement = ABS_MEASUREMENT_LABELS.get(label)

            if measurement is None:
                continue

            measurement_cells.append(
                ABSMeasurementCell(
                    row=row_index,
                    col=col_index,
                    measurement=cast(ABSMeasurement, measurement),
                    measurement_label=label,
                )
            )

    if not measurement_cells:
        raise ValueError("Could not identify any ABS measurement blocks.")

    measurement_cells.sort(key=lambda cell: (cell.row, cell.col))
    return measurement_cells


def find_abs_footer_start_row(raw_sheet: pd.DataFrame) -> Optional[int]:
    row_index = _last_nonblank_row(raw_sheet)

    if not _is_footer_row(raw_sheet.iloc[row_index]):
        return None

    footer_start = row_index

    for row_index in range(row_index - 1, -1, -1):
        row = raw_sheet.iloc[row_index]

        if count_nonempty_cells(row) == 0:
            break

        if not _is_footer_row(row):
            break

        footer_start = row_index

    return footer_start


def infer_abs_row_bounds(
    raw_sheet: pd.DataFrame,
    measurement_cells: list[ABSMeasurementCell],
    footer_start: Optional[int],
) -> ABSRowBounds:
    first_measurement_row = measurement_cells[0].row
    body_last = (
        footer_start - 1 if footer_start is not None else _last_nonblank_row(raw_sheet)
    )

    return ABSRowBounds(
        header_first=ABS_HEADER_FIRST_ROW_INDEX,
        header_last=first_measurement_row - 1,
        body_first=first_measurement_row,
        body_last=body_last,
        footer_start=footer_start,
    )


def infer_abs_subject_and_row_groups(
    raw_sheet: pd.DataFrame,
    row_bounds: ABSRowBounds,
    bounds: ABSSourceBounds,
    row_indents: Mapping[int, int],
) -> tuple[Optional[str], dict[int, list[str]], dict[int, int]]:
    subject_candidates: list[str] = []
    row_group_paths: dict[int, list[str]] = {}
    row_group_indents: dict[int, int] = {}

    current_path: list[str] = []
    current_group_indent: Optional[int] = None
    pending_subject_path: list[str] = []

    for row_index in range(row_bounds.body_first, row_bounds.body_last + 1):
        row_label = clean_cell_text(raw_sheet.iat[row_index, 0])
        if not row_label:
            continue

        row_indent = row_indents.get(row_index, 0)
        has_values = _row_has_measure_values(raw_sheet, row_index, bounds)

        if has_values:
            if current_path:
                row_group_paths[row_index] = current_path.copy()
                if current_group_indent is not None:
                    row_group_indents[row_index] = current_group_indent
            continue

        next_kind = _next_body_row_kind(
            raw_sheet,
            row_bounds,
            bounds,
            row_index + 1,
        )

        if next_kind == "label":
            pending_subject_path = [row_label]
            if row_label not in subject_candidates:
                subject_candidates.append(row_label)
        elif pending_subject_path:
            current_path = [*pending_subject_path, row_label]
            current_group_indent = row_indent
        else:
            current_path = [row_label]
            current_group_indent = row_indent

    subject = subject_candidates[0] if len(subject_candidates) == 1 else None
    return subject, row_group_paths, row_group_indents


def _header_label_for_column(
    raw_sheet: pd.DataFrame,
    header_row: int,
    col_index: int,
    first_col: int,
) -> str:
    for search_col in range(col_index, first_col - 1, -1):
        label = normalise_cell_text(raw_sheet.iat[header_row, search_col])
        if label:
            return label

    return ""


def flatten_abs_column_headers(
    raw_sheet: pd.DataFrame,
    row_bounds: ABSRowBounds,
    bounds: ABSSourceBounds,
) -> dict[int, list[str]]:
    column_headers: dict[int, list[str]] = {}

    for col_index in range(bounds.col_first, bounds.col_last + 1):
        labels: list[str] = []

        for row_index in range(row_bounds.header_first, row_bounds.header_last + 1):
            label = _header_label_for_column(
                raw_sheet,
                row_index,
                col_index,
                bounds.col_first,
            )

            if label and label not in labels:
                labels.append(label)

        column_headers[col_index] = labels

    return column_headers


def _build_record(
    table_number: int,
    table_title: str,
    measurement_cell: ABSMeasurementCell,
    subject: Optional[str],
    row_index: int,
    col_index: int,
    row_label: str,
    row_indent: int,
    row_group_path: list[str],
    row_group_indent: Optional[int],
    header_labels: list[str],
    value: object,
    value_text: str,
) -> dict[str, object]:
    value_marker = value_text.lower()
    record: dict[str, object] = {
        "table_number": table_number,
        "table_title": table_title,
        "measurement": measurement_cell.measurement,
        "measurement_label": measurement_cell.measurement_label,
        "subject": subject,
        "row_group": row_group_path[-1] if row_group_path else None,
        "row_group_path": " | ".join(row_group_path),
        "row_label": row_label,
        "row_indent": row_indent,
        "row_group_indent": row_group_indent,
        "column_header": " | ".join(header_labels),
        "source_row": row_index + 1,
        "source_column": col_index + 1,
        "value": value,
        "value_text": value_text,
        "is_suppressed": value_marker == "np",
        "is_not_available": value_marker == "na",
    }

    for header_index, header_label in enumerate(header_labels, start=1):
        record[f"column_header_{header_index}"] = header_label

    return record


def _build_parsed_subtable(
    raw_sheet: pd.DataFrame,
    table_number: int,
    table_title: str,
    measurement_cell: ABSMeasurementCell,
    subject: Optional[str],
    row_bounds: ABSRowBounds,
    bounds: ABSSourceBounds,
    row_indents: Mapping[int, int],
    row_group_paths: dict[int, list[str]],
    row_group_indents: dict[int, int],
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    column_headers = flatten_abs_column_headers(raw_sheet, row_bounds, bounds)

    for row_index in range(row_bounds.body_first, row_bounds.body_last + 1):
        row_label = clean_cell_text(raw_sheet.iat[row_index, 0])
        if not row_label:
            continue

        if not _row_has_measure_values(raw_sheet, row_index, bounds):
            continue

        row_group_path = row_group_paths.get(row_index, [])
        row_group_indent = row_group_indents.get(row_index)
        row_indent = row_indents.get(row_index, 0)

        for col_index in range(bounds.col_first, bounds.col_last + 1):
            value = raw_sheet.iat[row_index, col_index]
            value_text = clean_cell_text(value)

            if not value_text:
                continue

            records.append(
                _build_record(
                    table_number=table_number,
                    table_title=table_title,
                    measurement_cell=measurement_cell,
                    subject=subject,
                    row_index=row_index,
                    col_index=col_index,
                    row_label=row_label,
                    row_indent=row_indent,
                    row_group_path=row_group_path,
                    row_group_indent=row_group_indent,
                    header_labels=column_headers.get(col_index, []),
                    value=value,
                    value_text=value_text,
                )
            )

    return pd.DataFrame(records)


def _extract_raw_subtable(
    raw_sheet: pd.DataFrame, bounds: ABSSourceBounds
) -> pd.DataFrame:
    return raw_sheet.iloc[
        bounds.row_first : bounds.row_last + 1,
        bounds.col_first : bounds.col_last + 1,
    ].reset_index(drop=True)


def _build_subtable(
    raw_sheet: pd.DataFrame,
    table_number: int,
    table_title: str,
    sheet_rows: ABSRowBounds,
    measurement_cell: ABSMeasurementCell,
    measurement_cells: list[ABSMeasurementCell],
    row_indents: Mapping[int, int],
) -> ABSParsedTable:
    body_last = _infer_subtable_body_last(
        sheet_rows.body_last,
        measurement_cell,
        measurement_cells,
    )
    col_last = _infer_subtable_col_last(
        raw_sheet,
        sheet_rows,
        measurement_cell,
        measurement_cells,
        body_last,
    )

    bounds = ABSSourceBounds(
        row_first=measurement_cell.row,
        row_last=body_last,
        col_first=measurement_cell.col,
        col_last=col_last,
    )
    row_bounds = ABSRowBounds(
        header_first=sheet_rows.header_first,
        header_last=sheet_rows.header_last,
        body_first=measurement_cell.row + 1,
        body_last=body_last,
        footer_start=sheet_rows.footer_start,
    )

    subject, row_group_paths, row_group_indents = infer_abs_subject_and_row_groups(
        raw_sheet,
        row_bounds,
        bounds,
        row_indents,
    )

    return ABSParsedTable(
        value_kind=measurement_cell.measurement,
        source_measurement_label=measurement_cell.measurement_label,
        subject_label=subject,
        source_bounds=bounds,
        row_bounds=row_bounds,
        raw_table=_extract_raw_subtable(raw_sheet, bounds),
        parsed_table=_build_parsed_subtable(
            raw_sheet,
            table_number,
            table_title,
            measurement_cell,
            subject,
            row_bounds,
            bounds,
            row_indents,
            row_group_paths,
            row_group_indents,
        ),
    )


def extract_abs_subtables(
    raw_sheet: pd.DataFrame,
    table_number: int,
    table_title: str,
    rows: ABSRowBounds,
    measurement_cells: list[ABSMeasurementCell],
    row_indents: Mapping[int, int],
) -> list[ABSParsedTable]:
    subtables: list[ABSParsedTable] = []

    for measurement_cell in measurement_cells:
        subtables.append(
            _build_subtable(
                raw_sheet,
                table_number,
                table_title,
                rows,
                measurement_cell,
                measurement_cells,
                row_indents,
            )
        )

    return subtables


def extract_abs_metadata_sections(
    raw_sheet: pd.DataFrame,
    footer_start: Optional[int],
) -> Metadata:
    if footer_start is None:
        return {}

    footer_notes: list[str] = []

    for row_index in range(footer_start, len(raw_sheet)):
        row_texts = get_row_texts(raw_sheet.iloc[row_index])
        if row_texts:
            footer_notes.append(" | ".join(row_texts))

    return {"Footer": footer_notes}


def find_abs_table_sources(source_files: Mapping[str, str]) -> list[str]:
    table_sources: list[tuple[int, str]] = []

    for source_key, source_file in source_files.items():
        match = ABS_TABLE_SOURCE_KEY_PATTERN.fullmatch(source_key)
        if match is None:
            continue

        start_table = int(match.group("START"))
        table_sources.append((start_table, source_file))

    table_sources.sort()

    sorted_sources: list[str] = []
    for _, source_file in table_sources:
        sorted_sources.append(source_file)

    return sorted_sources


def list_abs_table_sheets(folder: Folder, source_file: str) -> list[str]:
    sheet_names = list_excel_sheets(folder, source_file)

    if not sheet_names:
        raise ValueError(f"ABS workbook has no sheets: {source_file}")

    first_sheet = sheet_names[0]
    if first_sheet != ABS_CONTENTS_SHEET_NAME:
        raise ValueError(
            f"Expected first ABS sheet to be {ABS_CONTENTS_SHEET_NAME!r} "
            f"in {source_file}, got {first_sheet!r}."
        )

    return sheet_names[1:]


def load_abs_row_indents(
    folder: Folder,
    source_file: str,
    sheet_name: str,
) -> dict[int, int]:
    file_path = resolve_spreadsheet_path(folder, source_file)
    workbook = load_workbook(file_path, data_only=True, read_only=False)

    try:
        worksheet = workbook[sheet_name]
        row_indents: dict[int, int] = {}

        for row_number in range(1, worksheet.max_row + 1):
            indent = worksheet.cell(row_number, 1).alignment.indent
            row_indents[row_number - 1] = int(indent or 0)

        return row_indents
    finally:
        workbook.close()
